"""Master Excel registry for regression test results.

Layout (sheet 'AllVersions'):
- Column A: Metric labels (one per row)
- For each version, a contiguous block of columns is added to the right:
  - Row 1: merged cells across the block containing the version name
  - Row 2: book names (one per column inside the block, order comes from files/ discovery)
  - Rows 3+: metric values (each row corresponds to a metric in column A)

This module uses openpyxl. Install with: pip install openpyxl
"""
from __future__ import annotations
import os
import json
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


MASTER_FILENAME = "master_results.xlsx"
SHEET_NAME = "AllVersions"

command_tags = ['document', 'figure', 'table','tabular' ,'itemize', 'enumerate', 'list','verbatim',
                'center', 'flushleft', 'flushright', 'mathequation', 'align' ,'quote',
                'equation', 'algorithm', 'algorithmic'    
                ]

# Metrics header list mirrored from regression runner
METRIC_ROWS = [
    "book_name",
    "scoring_method",
    "timestamp",
    "Score",
    "Percent Compiled",
    "Latex Errors",
    "Latex Warnings",
    "Bibtex Extracted %",
    "Bibtex Cited %",
    "Chapters %",
    "Sections %",
    "Subsections %",
    "Figures %",
    "Tables %",
    "Index Entries %",
    "------------",
    "Score Deductions:",
    "Latex Error % Deduction",
    "Latex Warning % Deduction",
    "Bibtex Extraction % Deduction",
    "Bibtex Citation % Deduction",
    "Chapters % Deduction",
    "Sections % Deduction",
    "Subsections % Deduction",
    "Figures % Deduction",
    "Tables % Deduction",
    "Index Entries % Deduction",
    "Total Begin-End Difference % Deduction",
    "-----",
    "Latex Errors Deduction",
    "Latex Warnings Deduction",
    "Bibtex Extraction Deduction",
    "Bibtex Citation Deduction",
    "Chapters Deduction",
    "Sections Deduction",
    "Subsections Deduction",
    "Figures Deduction",
    "Tables Deduction",
    "Index Entries Deduction",
    "Total Begin-End Difference Deduction",
    "------------",
    "Bibtex (metadata)",
    "Bibtex extracted (json)",
    "Entries Cited",
    "Bibtex (diff [meta - extracted])",
    "Bibtex (diff [extracted - cited])",
    "Chapters (metadata)",
    "Chapters (output)",
    "Chapters (diff [meta-out])",
    "Sections (metadata)",
    "Sections (output)",
    "Sections (diff [meta-out])",
    "Subsections (metadata)",
    "Subsections (output)",
    "Subsections (diff [meta-out])",
    "Figures (metadata)",
    "Figures (output)",
    "Included Graphics (output)",
    "Figures (diff [meta-out])",
    "Tables (metadata)",
    "Tables (output)",
    "Tables (diff [meta-out])",
    "Index Entries (metadata)",
    "Index Entries (output)",
    "Index Entries (diff [meta-out])",
    # begin-end and structural diagnostics added by multi-metric calculator
    "Total Begin",
    "Total End",
    "Total Begin-End Difference"
]
# per-command begin-end counts
for tag in command_tags:
    METRIC_ROWS.append(f"{tag.capitalize()} Begin")
    METRIC_ROWS.append(f"{tag.capitalize()} End")
    METRIC_ROWS.append(f"{tag.capitalize()} Difference")


def _master_path(results_dir: Path) -> Path:
    return results_dir / MASTER_FILENAME


def _ensure_workbook(path: Path, books: List[str]) -> None:
    """Create workbook with metric rows in column A if not exists."""
    if path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    # Column A - metric labels start at row 3 (we reserve rows 1-2 for version/book headers)
    ws.cell(row=2, column=1, value="Metric")
    for idx, metric in enumerate(METRIC_ROWS, start=3):
        ws.cell(row=idx, column=1, value=metric)

    wb.save(path)


def _load_or_create(results_dir: Path) -> Workbook:
    path = _master_path(results_dir)
    # Determine repository root relative to this module file (tests/regression_test/../..)
    repo_root = Path(__file__).resolve().parents[2]
    books = _discover_books(repo_root / "files")
    _ensure_workbook(path, books)
    return load_workbook(path)


def _discover_books(files_dir: Path) -> List[str]:
    """Discover books by scanning files/ directory (same logic as regression tester)."""
    if not files_dir.exists():
        return []
    books = []
    for item in sorted(files_dir.iterdir()):
        if item.is_dir():
            input_file = item / "inputs" / f"{item.name}.tex"
            output_file = item / "outputs" / f"{item.name}_final.tex"
            # include books even if output not present; keep consistent order
            books.append(item.name)
    return books


def _find_version_block(ws, version_name: str) -> Optional[tuple]:
    """Return (start_col_index, end_col_index) if a block for version exists."""
    # scan row 1 for merged cells or values
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        val = ws.cell(row=1, column=col).value
        if val == version_name:
            # walk left to find start of this block
            start = col
            while start > 2 and ws.cell(row=2, column=start - 1).value is not None:
                start -= 1
            # walk right to find block end
            end = col
            while end + 1 <= max_col and ws.cell(row=2, column=end + 1).value is not None:
                end += 1
            return start, end
    return None


def _add_version_block(ws, version_name: str, books: List[str]) -> int:
    """Append a block of columns for the version. Returns starting column index."""
    # Determine start column (append to the right)
    # Reserve column A for metrics; first block should start at column 2.
    # Leave a single empty separator column between existing content and the new block
    last_used = ws.max_column
    if last_used < 2:
        start_col = 2
    else:
        # place start two columns after last used to leave one empty column as separator
        start_col = last_used + 2

    # Add book columns and write book names in row 2
    for i, book in enumerate(books):
        col = start_col + i
        ws.cell(row=2, column=col, value=book)

    # Add an extra column for average score (label in row 2 only)
    avg_col = start_col + len(books)
    ws.cell(row=2, column=avg_col, value="avg.")

    # Merge top row across block (include avg column) and set version name
    end_col = avg_col if books else start_col
    merge_range = f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}1"
    ws.merge_cells(merge_range)
    ws.cell(row=1, column=start_col, value=version_name)
    ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center")

    return start_col


def _to_number(val):
    """Try to coerce a cell value to a float for comparison. Return None if not numeric."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return None
        # allow commas in numbers
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return None
    return None


def _ensure_avg_column(ws, start_col: int) -> int:
    """Ensure the version block starting at start_col has an 'avg.' column at its end.

    Returns the avg column index.
    This is safe to call for blocks created before the avg feature existed.
    """
    # scan row 2 from start_col to find the last non-empty header cell
    col = start_col
    max_col = ws.max_column
    # walk right while there is a book/label in row 2
    while col <= max_col and ws.cell(row=2, column=col).value is not None:
        col += 1
    last_filled = col - 1

    # if last_filled cell already says 'avg.' then return it
    if last_filled >= start_col and ws.cell(row=2, column=last_filled).value == "avg.":
        return last_filled

    # otherwise add avg at next column after last_filled
    avg_col = last_filled + 1
    ws.cell(row=2, column=avg_col, value="avg.")

    # update merged header for the version to include the new avg column
    # find the merged region start cell (row1,start_col) and extend merge to avg_col
    try:
        ws.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(avg_col)}1")
    except Exception:
        # fall back: set value and alignment on start_col
        pass

    return avg_col


def _color_master_by_version(ws):
    """Compare each version block with the previous block and color cells.

    Rules:
    - For 'Latex Errors' and 'Latex Warnings': increase -> red, decrease -> green
    - For other numeric metrics: increase -> green, decrease -> red
    - If values are equal or non-numeric or previous missing: no fill
    """
    max_col = ws.max_column
    # gather version blocks as (version_name, start, end)
    blocks = []
    seen_versions = set()
    for col in range(2, max_col + 1):
        v = ws.cell(row=1, column=col).value
        if v and v not in seen_versions:
            found = _find_version_block(ws, v)
            if found:
                start, end = found
                blocks.append((v, start, end))
                seen_versions.add(v)

    # sort by start column to ensure left-to-right chronological order
    blocks.sort(key=lambda x: x[1])

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # rows with metrics start at row 3
    metric_start = 3
    metric_end = metric_start + len(METRIC_ROWS) - 1

    negative_metrics = {"Latex Errors", "Latex Warnings"}

    # iterate over blocks and compare each to previous
    for i in range(1, len(blocks)):
        _, cur_start, cur_end = blocks[i]
        _, prev_start, prev_end = blocks[i - 1]

        # for each book column in current block
        for cur_col in range(cur_start, cur_end + 1):
            book = ws.cell(row=2, column=cur_col).value
            if not book:
                continue
            # find matching book in previous block
            prev_col = None
            for c in range(prev_start, prev_end + 1):
                if ws.cell(row=2, column=c).value == book:
                    prev_col = c
                    break
            if prev_col is None:
                continue

            # compare metric rows
            for row in range(metric_start, metric_end + 1):
                cell_cur = ws.cell(row=row, column=cur_col)
                cell_prev = ws.cell(row=row, column=prev_col)
                val_cur = _to_number(cell_cur.value)
                val_prev = _to_number(cell_prev.value)

                # only act on numeric comparisons where previous exists
                if val_cur is None or val_prev is None:
                    continue
                if val_cur == val_prev:
                    continue

                metric_label = ws.cell(row=row, column=1).value
                metric_label_lower = (metric_label or "").lower()
                # Errors and warnings: increase is bad
                if metric_label in negative_metrics:
                    if val_cur > val_prev:
                        cell_cur.fill = red_fill
                    else:
                        cell_cur.fill = green_fill
                # Difference metrics (contain 'diff' or 'difference'): smaller absolute diff is better
                elif "diff" in metric_label_lower or "difference" in metric_label_lower or "deduction" in metric_label_lower:
                    # compare absolute values: decrease in abs(diff) -> green
                    if abs(val_cur) < abs(val_prev):
                        cell_cur.fill = green_fill
                    else:
                        cell_cur.fill = red_fill
                else:
                    # for other numeric metrics: increase is good
                    if val_cur > val_prev:
                        cell_cur.fill = green_fill
                    else:
                        cell_cur.fill = red_fill


def update_master_excel(results_dir: Path, book_name: str, result: Dict) -> None:
    """Update (or create) the master Excel file with metrics for a single book and the current version."""
    wb = _load_or_create(results_dir)
    ws = wb[SHEET_NAME]

    # Discover files/ books using the repository root relative to this module
    repo_root = Path(__file__).resolve().parents[2]
    files_dir = repo_root / "files"
    books = _discover_books(files_dir)

    # Determine current version by reading version_control json (if available)
    version_file = repo_root / "codes" / "pdf_to_latex" / "version_control" / "version_history.json"
    version_name = "unknown"
    try:
        if version_file.exists():
            with open(version_file, "r") as f:
                versions = json.load(f)
            for v in versions:
                if v.get("is_current") and not v.get("is_deleted", False):
                    version_name = v.get("name")
                    break
    except Exception:
        version_name = "unknown"

    # Ensure metric rows exist in column A starting at row 3
    for idx, metric in enumerate(METRIC_ROWS, start=3):
        if ws.cell(row=idx, column=1).value is None:
            ws.cell(row=idx, column=1, value=metric)

    # Find or create version block
    block = _find_version_block(ws, version_name)
    if not block:
        start_col = _add_version_block(ws, version_name, books)
    else:
        start_col, end_col = block

    # Ensure avg column exists for this block (handles older workbooks)
    avg_col = _ensure_avg_column(ws, start_col)

    # Locate the column for the requested book within the version block
    # Recompute block boundaries for new workbook if needed
    if not block:
        # end_col includes the avg column
        end_col = start_col + len(books)

    try:
        book_index = books.index(book_name)
    except ValueError:
        # Book not found in discovered list: append it at the end of the current block
        book_index = len(books)
        books.append(book_name)
        # rewrite the book name row for the block so avg column stays at the end
        for i, b in enumerate(books):
            ws.cell(row=2, column=start_col + i, value=b)
        # write avg header after the last book
        avg_col = start_col + len(books)
        ws.cell(row=2, column=avg_col, value="avg.")
        # update merge to include new avg column
        new_end = avg_col
        ws.merge_cells(f"{get_column_letter(start_col)}1:{get_column_letter(new_end)}1")

    target_col = start_col + book_index

    # Fill metric rows
    # map result details/dict to metric names; prefer result['details'] when available
    details = result.get("details", {}) if isinstance(result.get("details", {}), dict) else {}

    # base fields
    mapping = {
        "book_name": result.get("book_name", book_name),
        "scoring_method": result.get("scoring_method", ""),
        "timestamp": result.get("timestamp", ""),
        "Score": result.get("score", 0.0),
    }

    # Merge details into mapping where key matches metric names
    for key in METRIC_ROWS:
        if key in details:
            mapping[key] = details.get(key)

    # Also allow keys from details to be written to any metric row with exact key match
    for row_idx, metric in enumerate(METRIC_ROWS, start=3):
        value = mapping.get(metric, "")
        ws.cell(row=row_idx, column=target_col, value=value)

    # Recompute average Score for this version block (avg column is just after last book)
    try:
        score_row_idx = 3 + METRIC_ROWS.index("Score")
    except ValueError:
        score_row_idx = None

    if score_row_idx is not None:
        # book score columns are start_col .. start_col+len(books)-1 (avg is at start_col+len(books))
        score_values = []
        for c in range(start_col, start_col + len(books)):
            v = _to_number(ws.cell(row=score_row_idx, column=c).value)
            if v is not None:
                score_values.append(v)
        avg_col = start_col + len(books)
        if score_values:
            avg_val = sum(score_values) / len(score_values)
            ws.cell(row=score_row_idx, column=avg_col, value=avg_val)
        else:
            # clear avg if no scores
            ws.cell(row=score_row_idx, column=avg_col, value=None)

    # Color cells by comparing versions, then save workbook
    try:
        _color_master_by_version(ws)
    except Exception:
        # best-effort coloring: don't fail the update if coloring errors
        pass

    wb.save(_master_path(results_dir))
